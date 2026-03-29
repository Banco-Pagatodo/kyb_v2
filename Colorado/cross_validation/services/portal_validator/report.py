"""
report.py — Generación de reportes Excel (.xlsx) y CSV
para los resultados de validación de portales gubernamentales.

Genera un archivo con las siguientes columnas por módulo:
  • Empresa / Razón Social
  • RFC
  • Identificador (clave elector, número de serie, RFC)
  • Estado (VIGENTE, VENCIDO, REVOCADO, NO_ENCONTRADO, …)
  • Detalle
  • Fecha de Consulta
  • Intentos
  • Datos Extra (JSON serializado)

Formatos soportados:
  - xlsx (requiere openpyxl)
  - csv  (sin dependencias extras)
"""
from __future__ import annotations

import csv
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Sequence

from .base import ResultadoPortal, logger

# ── Directorio de reportes ──
REPORT_DIR = Path(__file__).resolve().parents[2] / "reports"


def generar_reporte(
    resultados: Sequence[ResultadoPortal],
    *,
    formato: str = "xlsx",
    directorio: str | Path | None = None,
    nombre: str | None = None,
) -> Path:
    """
    Genera un reporte con los resultados de validación de portales.

    Args:
        resultados: Lista de ResultadoPortal
        formato: 'xlsx' o 'csv'
        directorio: Directorio de salida (default: Colorado/reports/)
        nombre: Nombre base del archivo (se agrega timestamp)

    Returns:
        Path al archivo generado
    """
    if not resultados:
        logger.warning("No hay resultados para generar reporte")
        return Path()

    # Directorio de salida
    out_dir = Path(directorio) if directorio else REPORT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)

    # Nombre del archivo
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = nombre or "validacion_portales"
    ext = "xlsx" if formato == "xlsx" else "csv"
    filepath = out_dir / f"{base_name}_{ts}.{ext}"

    # Preparar filas
    headers = [
        "Módulo",
        "Empresa",
        "RFC",
        "Identificador",
        "Estado",
        "Detalle",
        "Fecha Consulta",
        "Intentos",
        "Screenshot",
        "Datos Extra",
    ]

    rows = []
    for r in resultados:
        rows.append([
            r.modulo,
            r.empresa,
            r.rfc,
            r.identificador,
            r.estado.value if r.estado else "",
            r.detalle,
            r.fecha_consulta,
            str(r.intentos),
            r.screenshot or "",
            json.dumps(r.datos_extra, ensure_ascii=False) if r.datos_extra else "",
        ])

    if formato == "xlsx":
        filepath = _generar_xlsx(filepath, headers, rows, resultados)
    else:
        filepath = _generar_csv(filepath, headers, rows)

    logger.info(f"Reporte generado: {filepath}")
    return filepath


# ============================================================
#  Excel (.xlsx) con formato y colores
# ============================================================

def _generar_xlsx(
    filepath: Path,
    headers: list[str],
    rows: list[list[str]],
    resultados: Sequence[ResultadoPortal],
) -> Path:
    """Genera un archivo Excel con formato y colores."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning(
            "openpyxl no está instalado. Generando CSV como fallback. "
            "Instala con: pip install openpyxl"
        )
        csv_path = filepath.with_suffix(".csv")
        return _generar_csv(csv_path, headers, rows)

    wb = Workbook()

    # ── Hoja resumen ──
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    _escribir_resumen(ws_resumen, resultados)

    # ── Hoja de detalle ──
    ws_detalle = wb.create_sheet("Detalle")
    _escribir_detalle(ws_detalle, headers, rows)

    # ── Hojas por módulo ──
    modulos = sorted(set(r.modulo for r in resultados))
    for modulo in modulos:
        ws_mod = wb.create_sheet(modulo)
        mod_rows = [row for row, r in zip(rows, resultados) if r.modulo == modulo]
        _escribir_detalle(ws_mod, headers, mod_rows)

    wb.save(str(filepath))
    return filepath


def _escribir_resumen(ws, resultados: Sequence[ResultadoPortal]):
    """Escribe la hoja de resumen con estadísticas."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return

    # Título
    ws.merge_cells("A1:D1")
    ws["A1"] = "Reporte de Validación - Portales Gubernamentales"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E79")

    ws["A2"] = f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)

    ws["A3"] = f"Total registros: {len(resultados)}"

    # Estadísticas por módulo
    row = 5
    ws.cell(row=row, column=1, value="Módulo").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Total").font = Font(bold=True)
    ws.cell(row=row, column=3, value="Exitosos").font = Font(bold=True)
    ws.cell(row=row, column=4, value="Fallidos").font = Font(bold=True)
    ws.cell(row=row, column=5, value="% Éxito").font = Font(bold=True)

    modulos = sorted(set(r.modulo for r in resultados))
    row += 1
    for modulo in modulos:
        mod_results = [r for r in resultados if r.modulo == modulo]
        total = len(mod_results)
        exitosos = sum(
            1 for r in mod_results
            if r.estado and r.estado.value in (
                "ENCONTRADO", "VIGENTE", "VALIDO"
            )
        )
        fallidos = total - exitosos
        pct = (exitosos / total * 100) if total > 0 else 0

        ws.cell(row=row, column=1, value=modulo)
        ws.cell(row=row, column=2, value=total)
        ws.cell(row=row, column=3, value=exitosos)
        ws.cell(row=row, column=4, value=fallidos)
        ws.cell(row=row, column=5, value=f"{pct:.0f}%")

        # Color por porcentaje
        if pct >= 80:
            fill = PatternFill(start_color="C6EFCE", fill_type="solid")
        elif pct >= 50:
            fill = PatternFill(start_color="FFEB9C", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFC7CE", fill_type="solid")
        ws.cell(row=row, column=5).fill = fill

        row += 1

    # Estadísticas por estado
    row += 2
    ws.cell(row=row, column=1, value="Estado").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Cantidad").font = Font(bold=True)

    from collections import Counter
    estados = Counter(
        r.estado.value if r.estado else "DESCONOCIDO" for r in resultados
    )
    row += 1
    for estado, count in estados.most_common():
        ws.cell(row=row, column=1, value=estado)
        ws.cell(row=row, column=2, value=count)
        row += 1

    # Ajustar anchos
    for col_num in range(1, 6):
        ws.column_dimensions[
            chr(64 + col_num)
        ].width = 18


def _escribir_detalle(ws, headers: list[str], rows: list[list[str]]):
    """Escribe la hoja de detalle con encabezados coloreados."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    header_fill = PatternFill(start_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Estado → color
    COLOR_MAP = {
        "ENCONTRADO": "C6EFCE",
        "VIGENTE": "C6EFCE",
        "VALIDO": "C6EFCE",
        "NO_ENCONTRADO": "FFC7CE",
        "VENCIDO": "FFC7CE",
        "REVOCADO": "FFC7CE",
        "INVALIDO": "FFC7CE",
        "ERROR": "FFC7CE",
        "CAPTCHA_NO_RESUELTO": "FFEB9C",
        "SIN_DATOS": "D9E1F2",
    }

    # Encabezados
    for c, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Datos
    for r, row_data in enumerate(rows, 2):
        for c, value in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.border = thin_border

            # Color de estado
            if c == 5 and value in COLOR_MAP:
                cell.fill = PatternFill(
                    start_color=COLOR_MAP[value],
                    fill_type="solid",
                )

    # Ajustar anchos de columna
    column_widths = [12, 30, 15, 25, 18, 60, 20, 8, 40, 50]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Auto-filtro
    if rows:
        max_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{max_col}{len(rows) + 1}"


# ============================================================
#  CSV (fallback sin dependencias)
# ============================================================

def _generar_csv(
    filepath: Path,
    headers: list[str],
    rows: list[list[str]],
) -> Path:
    """Genera un archivo CSV."""
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)
    return filepath


# ============================================================
#  Utilidad: resumen rápido en consola
# ============================================================

def imprimir_resumen(resultados: Sequence[ResultadoPortal]) -> None:
    """Imprime un resumen rápido de los resultados en consola."""
    if not resultados:
        print("  (sin resultados)")
        return

    # Agrupar por módulo
    modulos: dict[str, list[ResultadoPortal]] = {}
    for r in resultados:
        modulos.setdefault(r.modulo, []).append(r)

    print("\n" + "=" * 70)
    print("  RESUMEN DE VALIDACIÓN — PORTALES GUBERNAMENTALES")
    print("=" * 70)

    for modulo, results in sorted(modulos.items()):
        total = len(results)
        ok = sum(
            1 for r in results
            if r.estado and r.estado.value in ("ENCONTRADO", "VIGENTE", "VALIDO")
        )
        fail = total - ok
        pct = (ok / total * 100) if total > 0 else 0
        bar = "█" * int(pct // 5) + "░" * (20 - int(pct // 5))

        print(f"\n  [{modulo}]  {bar}  {pct:.0f}% ({ok}/{total})")
        for r in results:
            icon = "✅" if r.estado and r.estado.value in (
                "ENCONTRADO", "VIGENTE", "VALIDO"
            ) else "❌"
            print(f"    {icon} {r.empresa:<30} {r.identificador:<25} → {r.estado.value if r.estado else '?'}")

    print("\n" + "=" * 70)
