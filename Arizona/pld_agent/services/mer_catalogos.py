"""
Catálogos de riesgo MER PLD/FT v7.0 — Banco PagaTodo.

Carga las tablas de riesgo del archivo Excel del modelo y las expone
como diccionarios estáticos para consulta rápida por el motor de cálculo.

Se cargan lazy (una sola vez al primer uso) para no penalizar el arranque
si el módulo no se utiliza.
"""
from __future__ import annotations

import logging
import re
import unicodedata
from pathlib import Path

logger = logging.getLogger("arizona.mer_catalogos")

_EXCEL_PATH = (
    Path(__file__).resolve().parent.parent
    / "docs"
    / "Modelo de riesgo de los clientes 2025.xlsx"
)

# ═══════════════════════════════════════════════════════════════════
#  Helpers de normalización
# ═══════════════════════════════════════════════════════════════════

def _normalizar(texto: str) -> str:
    """Minúsculas, sin acentos, sin espacios extra."""
    t = texto.strip().lower()
    t = unicodedata.normalize("NFD", t)
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    t = re.sub(r"\s+", " ", t)
    return t


# ═══════════════════════════════════════════════════════════════════
#  Catálogos estáticos (no dependen del Excel)
# ═══════════════════════════════════════════════════════════════════

# -- PESOS del modelo principal (hoja "Modelo") --
PESOS: dict[str, float] = {
    "tipo_persona":      0.10,
    "nacionalidad":      0.05,
    "antiguedad":        0.05,
    "actividad":         0.15,
    "ubicacion":         0.10,
    "producto":          0.05,
    "monto_recibido":    0.05,
    "monto_enviado":     0.05,
    "ops_recibidas":     0.05,
    "ops_enviadas":      0.05,
    "origen_recursos":   0.05,
    "destino_recursos":  0.05,
    "lpb":               0.10,
    "listas_negativas":  0.10,
    "pep":               0.10,
}

# -- Tipo de persona --
TIPO_PERSONA = {"pf": 1, "pfae": 2, "pm": 3}

# -- Productos y servicios --
PRODUCTOS: dict[str, int] = {
    "ya_ganaste":      1,
    "basica_nomina":   1,
    "adquirencia":     2,
    "fundadores":      2,
    "util":            3,
    "corporativa":     3,
}

# -- Países (lista negra / gris / GAFI) --
PAISES_LISTA_NEGRA: set[str] = {
    _normalizar(p) for p in [
        "República Democrática de Korea", "Irán", "Myanmar",
    ]
}
PAISES_LISTA_GRIS: set[str] = {
    _normalizar(p) for p in [
        "Algeria", "Angola", "Bolivia", "Bulgaria", "Burkina Faso",
        "Camerún", "Cote d'Ivoire", "República Democrática del Congo",
        "Haití", "Kenya", "Laos", "Líbano", "Mónaco", "Mozambique",
        "Namibia", "Nepal", "Nigeria", "Sudáfrica", "Sudán", "Syria",
        "Venezuela", "Vietnam", "Islas Vírgenes", "Yemen",
    ]
}

# -- Entidades federativas --
ENTIDAD_RIESGO: dict[str, int] = {
    _normalizar(e): v for e, v in {
        "Aguascalientes": 2,
        "Baja California": 3,
        "Baja California Sur": 1,
        "Campeche": 2,
        "Chiapas": 1,
        "Chihuahua": 3,
        "Ciudad de México": 3, "CDMX": 3,
        "Coahuila": 1,
        "Colima": 2,
        "Durango": 1,
        "Guanajuato": 2,
        "Guerrero": 2,
        "Hidalgo": 1,
        "Jalisco": 1,
        "México": 3, "Estado de México": 3,
        "Michoacan": 1, "Michoacán": 1,
        "Morelos": 3,
        "Nayarit": 1,
        "Nuevo León": 2, "Nuevo Leon": 2,
        "Oaxaca": 1,
        "Puebla": 2,
        "Queretaro": 2, "Querétaro": 2,
        "Quintana Roo": 2,
        "San Luis Potosi": 1, "San Luis Potosí": 1,
        "Sinaloa": 2,
        "Sonora": 2,
        "Tabasco": 2,
        "Tamaulipas": 1,
        "Tlaxcala": 1,
        "Veracruz": 1,
        "Yucatán": 1, "Yucatan": 1,
        "Zacatecas": 1,
    }.items()
}

# -- Alcaldías de CDMX (riesgo total, última columna del Excel) --
ALCALDIA_RIESGO: dict[str, float] = {
    _normalizar(a): v for a, v in {
        "Álvaro Obregón": 2.0,
        "Azcapotzalco": 1.5,
        "Benito Juárez": 1.5,
        "Coyoacán": 1.5,
        "Cuajimalpa de Morelos": 1.5,
        "Cuauhtémoc": 2.0,
        "Gustavo A. Madero": 2.5,
        "Iztacalco": 1.0,
        "Iztapalapa": 3.0,
        "La Magdalena Contreras": 2.0,
        "Miguel Hidalgo": 1.5,
        "Milpa Alta": 2.0,
        "Tláhuac": 2.0,
        "Tlalpan": 2.0,
        "Venustiano Carranza": 2.0,
        "Xochimilco": 2.0,
    }.items()
}

# -- Origen / Destino de recursos --
ORIGEN_DESTINO_RIESGO: dict[str, int] = {
    _normalizar(o): v for o, v in {
        "Pago de Nómina / Sueldos y Salarios": 1,
        "Nómina": 1, "Sueldos y Salarios": 1, "Nómina / Sueldos": 1,
        "Casa de Juegos y Apuestas (Casino)": 3, "Casino": 3,
        "Cheques de viajero": 3,
        "Crédito": 3,
        "Servicios de Inmuebles": 3, "Inmuebles": 3,
        "Comercio de Joyas, Metales y/o Piedras Preciosos, Relojes": 3,
        "Joyas": 3, "Joyería": 3, "Metales preciosos": 3,
        "Comercio de Obras de Arte": 3, "Obras de Arte": 3,
        "Comercio de Vehículos, Aéreos, Marítimos o Terrestres": 3,
        "Vehículos": 3,
        "Servicios de Blindaje de vehículos e inmuebles": 3, "Blindaje": 3,
        "Traslado de Custodia de dinero o valores": 3,
        "Fe Pública": 3, "Notaría": 3, "Notario": 3,
        "Donativos / Herencia": 3, "Donativos": 3, "Herencia": 3,
        "Aduanas": 3,
        "Arrendamiento de Inmuebles": 3, "Arrendamiento": 3,
        "Activos Virtuales": 3, "Criptomonedas": 3,
        "Otros": 2,
    }.items()
}

# -- Rangos PM-N4: volumen de operación --
RANGOS_MONTO_PM: list[tuple[float, float, int]] = [
    (1, 4_000_000, 1),
    (4_000_001, 20_000_000, 2),
    (20_000_001, float("inf"), 3),
]

# -- Rangos PM-N4: frecuencia de operación --
RANGOS_OPS_PM: list[tuple[int, int, int]] = [
    (1, 49_000, 1),
    (49_001, 300_000, 2),
    (300_001, 900_000_000, 3),  # "y más"
]

# -- Clasificación PM --
CLASIFICACION_PM: list[tuple[str, float, float]] = [
    ("BAJO",  85,  142),
    ("MEDIO", 143, 199),
    ("ALTO",  200, 255),
]


# ═══════════════════════════════════════════════════════════════════
#  Catálogo de actividades económicas (cargado lazy del Excel)
# ═══════════════════════════════════════════════════════════════════

_actividades: dict[str, int] | None = None
_actividades_por_codigo: dict[int, int] | None = None


def _cargar_actividades() -> None:
    """Lee la hoja 'Riesgo ActEconómica' del Excel y llena los dicts."""
    global _actividades, _actividades_por_codigo
    if _actividades is not None:
        return

    _actividades = {}
    _actividades_por_codigo = {}

    try:
        import openpyxl

        # Intentar abrir directamente; si falla por bloqueo (OneDrive/Excel),
        # copiar a temporal y leer desde ahí.
        excel_path = str(_EXCEL_PATH)
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
        except PermissionError:
            import shutil
            import tempfile
            tmp = Path(tempfile.gettempdir()) / "mer_actividades_tmp.xlsx"
            shutil.copy2(excel_path, str(tmp))
            wb = openpyxl.load_workbook(str(tmp), data_only=True, read_only=True)
            logger.info("Excel abierto desde copia temporal (archivo bloqueado).")

        ws = wb["Riesgo ActEconómica"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            # cols: idx, EconomicActivityTypeID, CNBVCode, Name, RiskValue
            if len(row) < 5 or row[3] is None:
                continue
            nombre = str(row[3]).strip()
            cnbv = row[2]
            valor = int(row[4]) if row[4] is not None else 2
            _actividades[_normalizar(nombre)] = valor
            if cnbv is not None:
                _actividades_por_codigo[int(cnbv)] = valor
        wb.close()
        logger.info("Catálogo actividades MER cargado: %d registros", len(_actividades))
    except Exception as exc:
        logger.error("No se pudo cargar catálogo de actividades: %s", exc)
        _actividades = {}
        _actividades_por_codigo = {}


def buscar_actividad(nombre_o_codigo: str) -> int | None:
    """
    Busca una actividad económica por nombre (parcial) o código CNBV.
    Retorna el valor de riesgo (1, 2 o 3) o None si no se encuentra.
    """
    _cargar_actividades()
    assert _actividades is not None and _actividades_por_codigo is not None

    # Intentar por código numérico
    try:
        codigo = int(nombre_o_codigo)
        if codigo in _actividades_por_codigo:
            return _actividades_por_codigo[codigo]
    except ValueError:
        pass

    # Búsqueda exacta normalizada
    clave = _normalizar(nombre_o_codigo)
    if clave in _actividades:
        return _actividades[clave]

    # Búsqueda parcial (contiene)
    for nombre_cat, valor in _actividades.items():
        if clave in nombre_cat or nombre_cat in clave:
            return valor

    return None


def obtener_actividades_riesgo_alto_medio() -> dict[str, list[str]]:
    """
    Retorna las actividades de Grupo 2 (medio) y Grupo 3 (alto) del catálogo
    CNBV para que el LLM pueda razonar sobre la clasificación.

    Solo son ~65 entradas (de 1,096 totales). Todo lo que no esté aquí
    es Grupo 1 (riesgo bajo) por defecto.
    """
    _cargar_actividades()
    assert _actividades is not None

    grupo_2: list[str] = []
    grupo_3: list[str] = []

    for nombre_norm, valor in _actividades.items():
        if valor == 3:
            grupo_3.append(nombre_norm)
        elif valor == 2:
            grupo_2.append(nombre_norm)

    return {"grupo_2": sorted(grupo_2), "grupo_3": sorted(grupo_3)}


# ═══════════════════════════════════════════════════════════════════
#  Funciones de lookup público
# ═══════════════════════════════════════════════════════════════════

def obtener_riesgo_pais(pais: str) -> int:
    """Retorna 300 (lista negra), 200 (lista gris) o 1 (GAFI/otros)."""
    clave = _normalizar(pais)
    if clave in PAISES_LISTA_NEGRA:
        return 300
    if clave in PAISES_LISTA_GRIS:
        return 200
    return 1


def obtener_riesgo_entidad(entidad: str) -> int:
    """Retorna 1, 2 o 3 según la zona de riesgo de la entidad federativa."""
    clave = _normalizar(entidad)
    if clave in ENTIDAD_RIESGO:
        return ENTIDAD_RIESGO[clave]
    # Búsqueda parcial
    for nombre_cat, valor in ENTIDAD_RIESGO.items():
        if clave in nombre_cat or nombre_cat in clave:
            return valor
    return 2  # Default: riesgo medio si no se encuentra


def obtener_riesgo_alcaldia(alcaldia: str) -> float:
    """
    Retorna el riesgo total de una alcaldía de CDMX.
    Si no se encuentra, retorna 2.0 (medio).
    """
    clave = _normalizar(alcaldia)
    if clave in ALCALDIA_RIESGO:
        return ALCALDIA_RIESGO[clave]
    for nombre_cat, valor in ALCALDIA_RIESGO.items():
        if clave in nombre_cat or nombre_cat in clave:
            return valor
    return 2.0


def obtener_riesgo_producto(producto: str) -> int:
    """Retorna 1, 2 o 3 según el producto contratado."""
    clave = _normalizar(producto)
    for nombre_cat, valor in PRODUCTOS.items():
        if _normalizar(nombre_cat) in clave or clave in _normalizar(nombre_cat):
            return valor
    return 3  # Default: riesgo alto si no reconoce el producto


def obtener_riesgo_monto_pm(monto: float) -> int:
    """Clasifica un monto (MXN) según tabla PM-N4."""
    for inf, sup, valor in RANGOS_MONTO_PM:
        if inf <= monto <= sup:
            return valor
    return 1


def obtener_riesgo_ops_pm(num_ops: int) -> int:
    """Clasifica frecuencia de operaciones según tabla PM-N4."""
    for inf, sup, valor in RANGOS_OPS_PM:
        if inf <= num_ops <= sup:
            return valor
    return 1


def obtener_riesgo_origen_destino(concepto: str) -> int:
    """Retorna 1, 2 o 3 según el catálogo de origen/destino de recursos."""
    clave = _normalizar(concepto)
    if clave in ORIGEN_DESTINO_RIESGO:
        return ORIGEN_DESTINO_RIESGO[clave]
    for nombre_cat, valor in ORIGEN_DESTINO_RIESGO.items():
        if clave in nombre_cat or nombre_cat in clave:
            return valor
    return 2  # Default: "Otros"


def clasificar_grado_riesgo_pm(puntaje: float) -> str:
    """Clasifica el puntaje total según la tabla PM."""
    for grado, inf, sup in CLASIFICACION_PM:
        if inf <= puntaje <= sup:
            return grado
    if puntaje > 255:
        return "ALTO"
    if puntaje < 85:
        return "BAJO"
    return "ALTO"
